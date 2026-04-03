import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---- SETTINGS ----
OUTPUT_FILE = f"correlation_{datetime.today().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
MIN_ROWS_REQUIRED = 30      # minimum non-null rows needed to attempt correlation
TOP_N_SCATTER = 500         # max rows to plot in scatter (keeps Excel fast)

# ---- STYLES ----
def style_header(cell, color="1F4E79"):
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def center_cell(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ---- STEP 1: Load CSV ----
def load_csv(path):
    df = pd.read_csv(path)
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    print(f"✅ Loaded {len(df)} rows.")
    print(f"   Numeric columns: {numeric_cols}")
    return df, numeric_cols

# ---- STEP 2: Audit nulls ----
def audit_nulls(df, numeric_cols):
    print("\n📋 Null audit:")
    audit = {}
    for col in numeric_cols:
        total = len(df)
        nulls = df[col].isnull().sum()
        valid = total - nulls
        pct = round((valid / total) * 100, 1)
        audit[col] = {"total": total, "valid": valid, "nulls": nulls, "pct_valid": pct}
        status = "✅" if pct >= 50 else "⚠️ "
        print(f"   {status} {col}: {valid} valid rows ({pct}%)")
    return audit

# ---- STEP 3: Correlation matrix ----
def compute_correlation(df, numeric_cols):
    # Only use columns with enough valid data
    usable = [c for c in numeric_cols if df[c].count() >= MIN_ROWS_REQUIRED]
    if len(usable) < 2:
        print(f"\n❌ Not enough usable columns for correlation (need at least 2 with {MIN_ROWS_REQUIRED}+ rows).")
        return None, usable

    corr = df[usable].corr(method="pearson")
    print(f"\n📊 Correlation matrix computed for: {usable}")
    return corr, usable

# ---- STEP 4: Interpret correlation strength ----
def interpret(val):
    abs_val = abs(val)
    if abs_val == 1.0:
        return "Perfect"
    elif abs_val >= 0.8:
        return "Very strong"
    elif abs_val >= 0.6:
        return "Strong"
    elif abs_val >= 0.4:
        return "Moderate"
    elif abs_val >= 0.2:
        return "Weak"
    else:
        return "Negligible"

def heatmap_color(val):
    # Red for negative, green for positive, white for zero
    if val > 0:
        intensity = int(val * 180)
        r = 255 - intensity
        g = 255
        b = 255 - intensity
    elif val < 0:
        intensity = int(abs(val) * 180)
        r = 255
        g = 255 - intensity
        b = 255 - intensity
    else:
        r, g, b = 255, 255, 255
    return f"{r:02X}{g:02X}{b:02X}"

# ---- STEP 5: Write Summary sheet ----
def write_summary(wb, df, numeric_cols, audit, input_file):
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Correlation Analysis Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Source: {input_file}  |  Rows: {len(df)}  |  Generated: {datetime.today().strftime('%d %B %Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Data quality section
    ws["A4"] = "DATA QUALITY AUDIT"
    ws["A4"].font = Font(bold=True, size=12, color="1F4E79")

    headers = ["Column", "Total Rows", "Valid Rows", "Null Rows", "% Valid", "Usable?"]
    for col_num, h in enumerate(headers, 1):
        style_header(ws.cell(row=5, column=col_num))
        ws.cell(row=5, column=col_num).value = h

    for i, (col, info) in enumerate(audit.items(), 6):
        usable = "✅ Yes" if info["pct_valid"] >= (MIN_ROWS_REQUIRED / len(df) * 100) else "⚠️ Low data"
        values = [col, info["total"], info["valid"], info["nulls"], f"{info['pct_valid']}%", usable]
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_num, value=val)
            cell.border = thin_border()
            center_cell(cell)
            # Colour % valid column
            if col_num == 5:
                pct = info["pct_valid"]
                if pct >= 70:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif pct >= 40:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    # Column widths
    for col_letter, width in zip(["A","B","C","D","E","F"], [22, 14, 14, 14, 12, 14]):
        ws.column_dimensions[col_letter].width = width

# ---- STEP 6: Write Correlation Matrix sheet ----
def write_matrix(wb, corr, usable_cols):
    ws = wb.create_sheet("Correlation Matrix")

    ws["A1"] = "Pearson Correlation Matrix"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    ws["A2"] = "Values range from -1 (perfect negative) to +1 (perfect positive). Diagonal is always 1.00."
    ws["A2"].font = Font(italic=True, color="888888", size=10)

    offset = 4  # start row for matrix

    # Header row and column
    for i, col in enumerate(usable_cols, 1):
        # Column headers
        cell = ws.cell(row=offset, column=i + 1, value=col)
        style_header(cell)
        # Row headers
        cell2 = ws.cell(row=offset + i, column=1, value=col)
        style_header(cell2)

    # Matrix values
    for row_i, row_col in enumerate(usable_cols, 1):
        for col_i, col_col in enumerate(usable_cols, 1):
            val = round(corr.loc[row_col, col_col], 3)
            cell = ws.cell(row=offset + row_i, column=col_i + 1, value=val)
            cell.number_format = "0.000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
            cell.fill = PatternFill("solid", fgColor=heatmap_color(val))
            # Bold the diagonal
            if row_col == col_col:
                cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 20
    for i in range(len(usable_cols)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

# ---- STEP 7: Write Pairs sheet (ranked correlation pairs) ----
def write_pairs(wb, corr, usable_cols):
    ws = wb.create_sheet("Ranked Pairs")

    ws["A1"] = "All Correlation Pairs — Ranked by Strength"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    headers = ["Column A", "Column B", "Correlation", "Strength", "Direction"]
    for col_num, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=col_num))
        ws.cell(row=3, column=col_num).value = h

    # Extract unique pairs (no duplicates, no diagonal)
    pairs = []
    cols = usable_cols
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            val = round(corr.loc[cols[i], cols[j]], 3)
            pairs.append((cols[i], cols[j], val))

    # Sort by absolute value descending
    pairs.sort(key=lambda x: abs(x[2]), reverse=True)

    for row_num, (col_a, col_b, val) in enumerate(pairs, 4):
        direction = "Positive ↑" if val > 0 else "Negative ↓" if val < 0 else "None"
        strength = interpret(val)
        row_vals = [col_a, col_b, val, strength, direction]

        for col_num, v in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=col_num, value=v)
            cell.border = thin_border()
            center_cell(cell)
            if col_num == 3:
                cell.number_format = "0.000"
                cell.fill = PatternFill("solid", fgColor=heatmap_color(val))
            # Highlight strong correlations
            if col_num == 4 and strength in ["Strong", "Very strong", "Perfect"]:
                cell.font = Font(bold=True, color="375623")
                cell.fill = PatternFill("solid", fgColor="C6EFCE")

    for col_letter, width in zip(["A","B","C","D","E"], [22, 22, 16, 16, 16]):
        ws.column_dimensions[col_letter].width = width

# ---- STEP 8: Write Scatter sheets for strongest pairs ----
def write_scatter(wb, df, pairs_data, top_n=3):
    # Pick top N strongest pairs (excluding perfect 1.0 diagonal)
    strong_pairs = [(a, b, v) for a, b, v in pairs_data if abs(v) < 1.0]
    strong_pairs = sorted(strong_pairs, key=lambda x: abs(x[2]), reverse=True)[:top_n]

    for col_a, col_b, val in strong_pairs:
        sheet_name = f"{col_a[:12]} vs {col_b[:12]}"
        ws = wb.create_sheet(title=sheet_name[:31])

        ws["A1"] = f"Scatter: {col_a} vs {col_b}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
        ws["A2"] = f"Pearson r = {val:.3f}  |  Strength: {interpret(val)}"
        ws["A2"].font = Font(italic=True, color="888888", size=10)

        # Write data (drop nulls for this pair, sample if large)
        pair_df = df[[col_a, col_b]].dropna()
        if len(pair_df) > TOP_N_SCATTER:
            pair_df = pair_df.sample(TOP_N_SCATTER, random_state=42)
            ws["A3"] = f"Note: Sampled {TOP_N_SCATTER} rows for performance."
            ws["A3"].font = Font(italic=True, color="888888", size=9)

        # Header
        ws.cell(row=4, column=1, value=col_a)
        ws.cell(row=4, column=2, value=col_b)
        style_header(ws.cell(row=4, column=1))
        style_header(ws.cell(row=4, column=2))

        for i, (_, row) in enumerate(pair_df.iterrows(), 5):
            ws.cell(row=i, column=1, value=round(float(row[col_a]), 4))
            ws.cell(row=i, column=2, value=round(float(row[col_b]), 4))

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18

        # Scatter chart
        num_rows = len(pair_df)
        chart = ScatterChart()
        chart.title = f"{col_a} vs {col_b}  (r={val:.3f})"
        chart.style = 10
        chart.x_axis.title = col_a
        chart.y_axis.title = col_b
        chart.width = 22
        chart.height = 14

        x_vals = Reference(ws, min_col=1, min_row=5, max_row=4 + num_rows)
        y_vals = Reference(ws, min_col=2, min_row=5, max_row=4 + num_rows)
        series = Series(y_vals, x_vals, title=f"r = {val:.3f}")
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)

        ws.add_chart(chart, "D4")
        print(f"   📈 Scatter: {col_a} vs {col_b} (r={val:.3f})")

# ---- MAIN ----
def main():
    print("🔗 Correlation Analysis\n")

    input_file = input("📂 Enter CSV filename (e.g. data.csv): ").strip()

    try:
        pd.read_csv(input_file, nrows=0)
    except FileNotFoundError:
        print(f"\n❌ File not found: '{input_file}'")
        print("   Make sure the file is in the same folder as this script.")
        return

    df, numeric_cols = load_csv(input_file)

    if len(numeric_cols) < 2:
        print("\n❌ Need at least 2 numeric columns for correlation analysis.")
        return

    # Audit nulls
    audit = audit_nulls(df, numeric_cols)

    # Compute correlation
    corr, usable_cols = compute_correlation(df, numeric_cols)
    if corr is None:
        return

    # Extract pairs for scatter sheets
    pairs_data = []
    for i in range(len(usable_cols)):
        for j in range(i + 1, len(usable_cols)):
            val = round(corr.loc[usable_cols[i], usable_cols[j]], 3)
            pairs_data.append((usable_cols[i], usable_cols[j], val))

    # Write Excel
    print("\n📝 Writing report...")
    wb = Workbook()
    write_summary(wb, df, numeric_cols, audit, input_file)
    write_matrix(wb, corr, usable_cols)
    write_pairs(wb, corr, usable_cols)
    write_scatter(wb, df, pairs_data, top_n=3)

    wb.save(OUTPUT_FILE)
    print(f"\n💾 Saved: {OUTPUT_FILE}")
    print(f"✅ Done! Check each sheet for results.")

main()